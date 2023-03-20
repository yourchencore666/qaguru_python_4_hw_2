import csv, zipfile, openpyxl
from PyPDF2 import PdfReader
from openpyxl.reader.excel import load_workbook


def test_create_zipfile(create_zip, delete_zip):
    assert zipfile.is_zipfile('resources/archive.zip') == True


def test_check_pdf_file_in_zip(create_zip, delete_zip):
    with open("resources/33.pdf", "rb") as pdf_file:
        reader = PdfReader(pdf_file)
        page_count = len(reader.pages)
        pdf_text = reader.pages[0].extract_text()

    with zipfile.ZipFile('resources/archive.zip', 'r') as zip_file:
        with zip_file.open("33.pdf", "r") as pdf_file:
            reader = PdfReader(pdf_file)
            page_count_zip = len(reader.pages)
            pdf_text_zip = reader.pages[0].extract_text()

    assert pdf_text == pdf_text_zip
    assert page_count == page_count_zip


def test_check_xls_file_in_zip(create_zip, delete_zip):
    file = load_workbook('resources/file_example_XLSX_10.xlsx')

    with open('resources/file_example_XLSX_10.xlsx', 'r') as xls_file:
        sheet = file.active
        xlsx_rows = sheet.max_row

    with zipfile.ZipFile('resources/archive.zip', 'r') as zip_file:
        xls_file = zip_file.open('file_example_XLSX_10.xlsx')
        xlsx_reader = openpyxl.load_workbook(xls_file)
        sheet = xlsx_reader.active
        xlsx_rows_zip = sheet.max_row

    assert xlsx_rows == xlsx_rows_zip


def test_check_csv_file_in_zip(create_zip, delete_zip):
    row_count_zip = 0

    with open('resources/spam3.csv', "r") as csv_file:
        file = csv.reader(csv_file, delimiter=";")
        row_count = sum(1 for _ in file)

    with zipfile.ZipFile('resources/archive.zip', 'r') as zip_file:
        with zip_file.open('spam3.csv', "r") as csv_file_zip:
            for _ in csv_file_zip:
                row_count_zip += 1

    assert row_count == row_count_zip
